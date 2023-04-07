# import json

  
# # Opening JSON file
# f = open('test.json')
  
# # returns JSON object as 
# # a dictionary
# data = json.load(f)
# # Closing file
# f.close()
# # Iterating through the json
# # list
# for i in data:

#     print(i['SKU_SHORT_DESCRIPTION'])
#     print("jewelry".lower() in i['SKU_SHORT_DESCRIPTION'].lower())
  

# import openpyxl
 
# # Define variable to load the dataframe
# dataframe = openpyxl.load_workbook("NER_Attributes.xlsx")
 
# # Define variable to read sheet
# dataframe1 = dataframe.active
 
# # Iterate the loop to read the cell values
# for row in range(0, dataframe1.max_row):
#     for col in dataframe1.iter_cols(1, dataframe1.max_column):
#         print(col[row].value)


from openpyxl import load_workbook
workbook = load_workbook('NER_Attributes.xlsx')
sheet = workbook.active
row_count = sheet.max_row
for i in range(row_count):
   print(sheet.cell(row=i+1, column=2).value)
   for j in range(1,12):
      print(j)


